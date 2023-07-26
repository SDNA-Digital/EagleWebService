from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import psycopg2 as pg
from django.conf import settings

def InportExcel(request, caminho, id):

    html = "<html><body>Done! %s </body></html>"

    conn = pg.connect(
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        host=settings.DB_HOST,
        port=settings.DB_PORT,
        database=settings.DB_NAME
	)

    cursor = conn.cursor()
    query = f"""select ConfigArquivosTabela from ConfigArquivos where ConfigArquivosId = {id}"""

    cursor.execute(query)
    resultado = cursor.fetchone()

    
    Tabela = resultado[0]

    query = f"""select ConfigArquivosAssociacaoAtt, ConfigArquivosAssociacaoColuna from ConfigArquivosAssociacao where ConfigArquivosId = {id} order by configarquivosassociacaocoluna"""
    cursor.execute(query)
    resultados = cursor.fetchall()
    contador = 0
    Att = []
    Coluna = []
    for resultado in resultados:
        Att.append(resultado[0])
        Coluna.append(resultado[1])
        contador += 1

    listaatt = ''
    contadoraux = 0
    for attini in Att:
        if contadoraux == contador - 1:
            listaatt += attini
            break
        else:
            contadoraux += 1
            listaatt += attini + ','
    
    # abre o arquivo excel
    wb = load_workbook(caminho)
    ws = wb.active

    # itera sobre todas as linhas no arquivo Excel
    for row in ws.iter_rows(min_row=2):
        valores = []

        # itera sobre todas as colunas na linha atual
        for i in range(contador):
            j = i + 1
            letter = get_column_letter(j)
            valores.append(ws[letter + str(row[0].row)].value)
            formato = ws[letter + str(row[0].row)].value

        contadoraux = 0
        listavalor = ''
        for valorcampo in valores:
            tipo = str(type(valorcampo))
            tipocompara = "<class 'datetime.datetime'>"
            concatena = valorcampo

            if tipo == tipocompara:
                concatena = ''
                concatena = str(valorcampo)
                concatena = "to_date('" + concatena[8:10] + "/" + concatena[5:7] + "/" + concatena[0:4] + "', 'DD/MM/YYYY')"

            if contadoraux == contador - 1:
                if tipo == tipocompara:
                    listavalor += str(concatena)
                else:
                    if concatena is not None:
                        listavalor += "'" + str(concatena) + "'"
                    break
            else:
                if tipo == tipocompara:
                    listavalor += str(concatena) + ','
                else:
                    listavalor += "'" + str(concatena) + "'" + ','
                contadoraux += 1

        
        montaquery = 'INSERT INTO ' + Tabela + ' (' + listaatt + ') Values(' + listavalor + ')'
        query = montaquery
       
        cursor.execute(query)
        conn.commit()

    return HttpResponse(html)

def ExportExcel(request, caminho, id):

    conn = pg.connect(
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        host=settings.DB_HOST,
        port=settings.DB_PORT,
        database=settings.DB_NAME
	)

    cursor = conn.cursor()
    query = f"""select ConfigArquivosTabela from ConfigArquivos where ConfigArquivosId = {id}"""
    cursor.execute(query)

    confregistro = cursor.fetchone()

    tabela = confregistro[0]

    query = f"""select * from {tabela}"""

    cursor.execute(query)
    registros = cursor.fetchall()


    wb = Workbook()
    ws = wb.active


    cabecalhos = [i[0] for i in cursor.description]
    for coluna, cabecalho in enumerate(cabecalhos):
        ws.cell(row=1, column=coluna+1, value=cabecalho)


    for linha, registro in enumerate(registros):
        for coluna, valor in enumerate(registro):
            ws.cell(row=linha+2, column=coluna+1, value=valor)

    
    wb.save(caminho)
  
    return  HttpResponse('Done')
